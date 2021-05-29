import ast
import glob
import os
from pathlib import Path

import openpyxl as ox
import datetime


def file_open(file_path):
    """
    text fileを開いて辞書型に書き換える
    return {}
    """
    with open(file_path, mode="r", encoding="utf-8") as f:
        data = f.read()
    data = ast.literal_eval(data)
    return data

def remove(file_path):
    """
    ファイルを消す
    """
    os.remove(file_path)

def get_logs(folder_name,remove_file):
    """
    一週間分のテキストを1日づつ呼び出す
    返り値 {day:{{id:{name:'...','time':[... , ...]}}}, ...}
    """
    day_file_list=sorted(glob.glob(folder_name+"/*.txt"))
    data_dict={}
    for file_path in day_file_list:
        data = file_open(file_path)
        if remove_file==True:
            remove(file_path)
        file_path = os.path.basename(file_path).split(".")[0][4:]
        data_dict.update([(file_path,data)])
    return data_dict

def get_term():
    """
    呼び出した日の前日を含む一週間を表示
    """
    now = datetime.datetime.now()
    end = now + datetime.timedelta(hours=12) - datetime.timedelta(days=1)
    start = end - datetime.timedelta(days=7)
    term = start.strftime("%Y%m%d")+"-"+end.strftime("%m%d")
    return term


def set_sheet(template,output_xlsx,log_dict,user_dict={}):
    #template.xlsxを呼び出し、log_dictに記述しているデータを上書きする。そのあと、output_xlsxにセーブ
    term = get_term()
    output_name = output_xlsx+"/"+term+"_研究室入退室管理簿_計算生物学研究室.xlsx"
    if Path(output_name).exists():
        Path(output_name).rename(output_name + '.bak_' + datetime.datetime.now().strftime('%Y%m%d_%H%M%S'))
    original = ox.load_workbook(template)
    sheet = original['Sheet1']
    index_number = 0
    for day in log_dict.keys():
        #一日分を持ってくる
        for user_id in log_dict[day].keys():
            if user_id in user_dict.keys():
                name = user_dict[user_id]['name']
            else:
                name = log_dict[day][user_id]["name"]

            if len(log_dict[day][user_id]['time'])==1:
                #退出記録がない場合は"22:00"を追加
                log_dict[day][user_id]['time'].append("22:00")

            sheet.cell(row=5+index_number, column=2, value=day[:2]+"/"+day[2:])#日付
            sheet.cell(row=5+index_number, column=3, value=int(user_id))#student_ID　学生番号
            sheet.cell(row=5+index_number, column=4, value=name)#名前
            sheet.cell(row=5+index_number, column=5, value="クリエーションコア3階 遠里研究室")#場所
            sheet.cell(row=5+index_number, column=6, value=log_dict[day][user_id]['time'][0])#入出時間
            sheet.cell(row=5+index_number, column=7, value=log_dict[day][user_id]['time'][1])#退出時間
            index_number+=1
    original.save(output_name)
    return output_name


def create_xlsx(template,log_folder,output_xlsx,remove_file=False,user_dict={}):
    """
    main関数でこれを呼び出す
    xlsxファイルを作り出す
    return 作り出したxlsxファイルの名前
    """
    os.makedirs(output_xlsx, exist_ok=True)
    log_dict = get_logs(log_folder,remove_file)
    output_name = set_sheet(template,output_xlsx,log_dict,user_dict)
    return output_name

def main():
    template="../template/template.xlsx"
    log_folder="../log"
    output_xlsx="../output"
    remove_file=False
    user_dict={111:{'name': 'hogehoge', 'slack_ID': '...'}}
    output_name = create_xlsx(template,log_folder,output_xlsx,remove_file,user_dict)


if  __name__=="__main__":
    main()
